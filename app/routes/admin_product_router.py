import io
import re
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, File, HTTPException, Query, UploadFile
from fastapi.responses import HTMLResponse
from openpyxl import load_workbook  # type: ignore
from pydantic import BaseModel

from app.services.product_sync_service import product_sync_manager
from app.database.connection import database_connection
from app.config import IMAGE_EMBEDDING_MODEL, IMAGE_EMBEDDING_PRETRAINED


router = APIRouter(prefix="/admin/products", tags=["Product Admin"])
PAGE_PATH = Path(__file__).resolve().parent.parent / "static" / "product_admin.html"
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
SKU_HEADERS = {"sku", "ma san pham", "mã sản phẩm", "product code", "product_code"}


class SkuImportRequest(BaseModel):
    skus: list[str]


def _start_job(skus: list[str], background_tasks: BackgroundTasks) -> dict:
    try:
        job = product_sync_manager.create(skus)
    except ValueError as error:
        raise HTTPException(status_code=422, detail=str(error)) from error
    background_tasks.add_task(product_sync_manager.run, job.id)
    return job.public()


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().casefold())


def _read_excel_skus(content: bytes) -> list[str]:
    try:
        workbook = load_workbook(
            filename=io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as error:
        raise ValueError("File Excel không hợp lệ hoặc bị hỏng.") from error

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise ValueError("File Excel không có dữ liệu.")

    headers = [_normalize_header(value) for value in header_row]
    sku_index = next(
        (index for index, header in enumerate(headers) if header in SKU_HEADERS),
        None,
    )
    if sku_index is None:
        raise ValueError(
            "Excel phải có cột SKU, Mã sản phẩm hoặc Product Code."
        )

    skus: list[str] = []
    for row in rows:
        if sku_index >= len(row):
            continue
        value = row[sku_index]
        if value is not None and str(value).strip():
            skus.append(str(value).strip())

    if not skus:
        raise ValueError("Cột SKU không có mã sản phẩm.")
    return skus


@router.get("", response_class=HTMLResponse)
def product_admin_page() -> HTMLResponse:
    return HTMLResponse(
        PAGE_PATH.read_text(encoding="utf-8"),
        headers={"Cache-Control": "no-store"},
    )


@router.post("/api/import-skus")
def import_skus(
    data: SkuImportRequest,
    background_tasks: BackgroundTasks,
) -> dict:
    return _start_job(data.skus, background_tasks)


@router.post("/api/import-excel")
async def import_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
) -> dict:
    suffix = Path(file.filename or "").suffix.casefold()
    if suffix != ".xlsx":
        raise HTTPException(status_code=415, detail="Chỉ hỗ trợ file .xlsx.")

    content = await file.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File vượt quá 5 MB.")

    try:
        skus = _read_excel_skus(content)
    except ValueError as error:
        raise HTTPException(status_code=422, detail=str(error)) from error
    return _start_job(skus, background_tasks)


@router.get("/api/catalog")
def synchronized_products(
    search: str = Query(default="", max_length=100),
    limit: int = Query(default=20, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
) -> dict:
    keyword = search.strip()
    where = ""
    search_parameters: list[object] = []
    if keyword:
        where = """
            WHERE p.product_code ILIKE %s
               OR p.title ILIKE %s
               OR p.product_type ILIKE %s
        """
        pattern = f"%{keyword}%"
        search_parameters = [pattern, pattern, pattern]

    with database_connection() as connection:
        total = connection.execute(
            f"SELECT COUNT(*) AS count FROM products p {where}",
            search_parameters,
        ).fetchone()["count"]

        rows = connection.execute(
            f"""
            SELECT
                p.product_code,
                p.title,
                p.product_type,
                p.status,
                p.updated_at,
                COUNT(DISTINCT pv.id) AS variant_count,
                COUNT(DISTINCT pi.id) FILTER (
                    WHERE pi.is_active = TRUE
                ) AS image_count,
                COUNT(DISTINCT pi.id) FILTER (
                    WHERE pi.is_active = TRUE
                      AND COALESCE(pi.local_path, '') <> ''
                ) AS local_image_count,
                COUNT(DISTINCT pie.product_image_id) AS embedding_count,
                STRING_AGG(DISTINCT pc.color, ', ' ORDER BY pc.color)
                    AS colors
            FROM products p
            LEFT JOIN product_variants pv ON pv.product_id = p.id
            LEFT JOIN product_images pi ON pi.product_id = p.id
            LEFT JOIN product_image_embeddings pie
                ON pie.product_image_id = pi.id
               AND pie.model_name = %s
               AND pie.pretrained_name = %s
            LEFT JOIN product_colors pc ON pc.product_id = p.id
            {where}
            GROUP BY p.id
            ORDER BY p.updated_at DESC, p.product_code
            LIMIT %s OFFSET %s
            """,
            [
                IMAGE_EMBEDDING_MODEL,
                IMAGE_EMBEDDING_PRETRAINED,
                *search_parameters,
                limit,
                offset,
            ],
        ).fetchall()

    products = []
    for row in rows:
        item = dict(row)
        local_count = int(item["local_image_count"] or 0)
        embedding_count = int(item["embedding_count"] or 0)
        item["ai_ready"] = (
            item["status"] == "ACTIVE"
            and local_count > 0
            and embedding_count >= local_count
        )
        item["updated_at"] = (
            item["updated_at"].isoformat()
            if item.get("updated_at")
            else None
        )
        products.append(item)

    return {
        "total": int(total),
        "limit": limit,
        "offset": offset,
        "embedding_model": IMAGE_EMBEDDING_MODEL,
        "products": products,
    }


@router.get("/api/jobs/{job_id}")
def get_job(job_id: str) -> dict:
    job = product_sync_manager.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Không tìm thấy tác vụ.")
    return job
